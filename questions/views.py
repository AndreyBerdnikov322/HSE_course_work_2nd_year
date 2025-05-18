from django.forms import formset_factory
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import render_to_string
from .models import Question, QuestionType, Image, Answer
from .forms import QuestionForm, ImageForm
from django.forms import inlineformset_factory
from django.views.generic import UpdateView
from django.contrib.auth import login, authenticate
from django.contrib.auth.views import LoginView, LogoutView
from .forms import RegisterForm, LoginForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django import forms
from django.forms import inlineformset_factory
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
import openpyxl
import os
from openpyxl.drawing.image import Image as ExcelImage
from django.conf import settings
from openpyxl.styles import Border, Side, Alignment
# Create your views here.

class QuestionListView(LoginRequiredMixin, ListView):
    model = Question
    template_name = 'questions/question_list.html'
    context_object_name = 'questions'
    login_url = '/login/'

@login_required
def question_detail(request, pk):
    question = get_object_or_404(Question, pk=pk)
    return render(request, 'questions/question_detail.html', {'question': question})

class QuestionCreateView(CreateView):
    model = Question
    form_class = QuestionForm
    template_name = 'questions/question_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        AnswerFormSet = inlineformset_factory(
            Question, 
            Answer, 
            fields=('text', 'is_correct'), 
            extra=1,
            widgets={
                'is_correct': forms.CheckboxInput(attrs={
                    'class': 'form-check-input', 
                    'type': 'checkbox',
                    'data-switch': 'true'
                }),
                'text': forms.TextInput(attrs={'class': 'form-control'})
            }
        )
        context['answer_formset'] = AnswerFormSet(self.request.POST or None)
        context['existing_images'] = []
        context['question_list_url'] = reverse('question_list')
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        answer_formset = context['answer_formset']
        
        if answer_formset.is_valid():
            self.object = form.save()
            answer_formset.instance = self.object
            answer_formset.save()
            
            # Обработка изображений
            images = self.request.FILES.getlist('new_images')
            captions = self.request.POST.getlist('new_captions')
            for image, caption in zip(images, captions):
                if image:
                    Image.objects.create(
                        question=self.object,
                        image=image,
                        caption=caption
                    )
            
            return redirect(self.get_success_url())
        return self.render_to_response(self.get_context_data(form=form))

class CustomLoginView(LoginView):
    form_class = LoginForm
    template_name = 'registration/login.html'
    redirect_authenticated_user = True 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Авторизация'
        if 'register_form' not in context:
            context['register_form'] = RegisterForm()
        return context


def register_view(request):
    if request.user.is_authenticated:
        return redirect('question_list')  # Редирект для авторизованных

    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)  # Автоматический вход после регистрации
            return redirect('question_list')
    else:
        form = RegisterForm()

    return render(request, 'registration/register.html', {'form': form})


class CustomLogoutView(LogoutView):
    next_page = 'login'

    @method_decorator(csrf_exempt)
    @method_decorator(require_POST)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

class QuestionUpdateView(UpdateView):
    model = Question
    form_class = QuestionForm
    template_name = 'questions/question_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        AnswerFormSet = inlineformset_factory(
            Question, 
            Answer, 
            fields=('text', 'is_correct'), 
            extra=0,
            can_delete=True,
            widgets={
                'is_correct': forms.CheckboxInput(attrs={
                    'class': 'form-check-input', 
                    'type': 'checkbox',
                    'data-switch': 'true'
                }),
                'text': forms.TextInput(attrs={'class': 'form-control'})
            }
        )
        
        if self.request.POST:
            context['answer_formset'] = AnswerFormSet(
                self.request.POST, 
                instance=self.object,
                prefix='answers'
            )
        else:
            context['answer_formset'] = AnswerFormSet(
                instance=self.object,
                prefix='answers'
            )
            
        context['existing_images'] = self.object.images.all()
        context['question_list_url'] = reverse('question_list')
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        answer_formset = context['answer_formset']
        
        if form.is_valid() and answer_formset.is_valid():
            self.object = form.save()
            
            # Сохраняем ответы с учетом удаленных
            instances = answer_formset.save(commit=False)
            for instance in instances:
                instance.save()
            for obj in answer_formset.deleted_objects:
                obj.delete()
            
            # Обработка удаленных изображений
            deleted_images = self.request.POST.getlist('delete_images')
            Image.objects.filter(pk__in=deleted_images).delete()
            
            # Добавление новых изображений
            images = self.request.FILES.getlist('new_images')
            captions = self.request.POST.getlist('new_captions')
            for image, caption in zip(images, captions):
                if image:
                    Image.objects.create(
                        question=self.object,
                        image=image,
                        caption=caption
                    )
            
            return super().form_valid(form)
        return self.render_to_response(self.get_context_data(form=form))

@login_required
@require_http_methods(["POST"])
def question_delete(request, pk):
    question = get_object_or_404(Question, pk=pk)
    question.delete()
    return redirect('question_list')

@require_POST
@login_required
def delete_answer(request, pk):
    answer = get_object_or_404(Answer, pk=pk)
    answer.delete()
    return JsonResponse({'success': True})

def add_image(request, question_pk):
    question = get_object_or_404(Question, pk=question_pk)
    if request.method == 'POST':
        images = request.FILES.getlist('new_images')
        captions = request.POST.getlist('new_captions')
        
        for image, caption in zip(images, captions):
            if image:
                if image.size > 5*1024*1024:
                    messages.error(request, f"Файл {image.name} слишком большой (макс. 5MB)")
                    continue
                
                Image.objects.create(
                    question=question,
                    image=image,
                    caption=caption
                )
        
        if not images:
            messages.error(request, "Необходимо выбрать хотя бы один файл")
        else:
            messages.success(request, "Изображения успешно добавлены!")
        
        return redirect('question_update', pk=question.pk)
    
    return HttpResponse(status=405)  # Method Not Allowed

@login_required
def delete_image(request, pk):
    image = get_object_or_404(Image, pk=pk)
    if request.method == 'POST':
        image.delete()
        return HttpResponse(status=204, headers={'HX-Trigger': 'imageListChanged'})
    return render(request, 'questions/image_confirm_delete.html', {'image': image})

@login_required
def image_list(request, question_pk):
    question = get_object_or_404(Question, pk=question_pk)
    images = question.images.all()
    return render(request, 'questions/image_list.html', {'images': images, 'question': question})

@login_required
def add_image_form(request):
    """Возвращает HTML-форму для добавления изображения"""
    return render(request, 'questions/partials/image_form.html')

@login_required
def question_create(request):
    if request.method == 'POST':
        form = QuestionForm(request.POST, request.FILES)
        if form.is_valid():
            question = form.save()
            
            # Обработка изображений
            for image, caption in zip(
                request.FILES.getlist('images'),
                request.POST.getlist('image_captions')
            ):
                if image:  # Проверяем, что файл был загружен
                    Image.objects.create(
                        question=question,
                        image=image,
                        caption=caption
                    )
            
            return redirect('question_detail', pk=question.pk)
    else:
        form = QuestionForm()
    
    question_types = QuestionType.objects.all()
    return render(request, 'questions/question_form.html', {
        'form': form,
        'question_types': question_types
    })

@login_required
def settings_view(request):
    return render(request, 'questions/settings.html')

@login_required
@require_POST
def delete_account(request):
    user = request.user
    if user.check_password(request.POST.get('password', '')):
        user.delete()
        logout(request)
        messages.success(request, "Ваш аккаунт был успешно удален.")
        return redirect('question_list')
    else:
        messages.error(request, "Неверный пароль. Аккаунт не удален.")
        return redirect('settings')

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Ваш пароль был успешно изменен!')
            return redirect('settings')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    return redirect('settings')

@login_required
def export_questions_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    base_headers = ["ID", "Текст вопроса", "Тип", "Ответы", "Правильные ответы"]
    max_images = max([q.images.count() for q in Question.objects.all()] + [0])
    image_headers = [f"Изображение {i+1}" for i in range(max_images)]
    headers = base_headers + image_headers
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2

    for question in Question.objects.all():
        answers = question.answers.all()
        all_answers = "; ".join([a.text for a in answers])
        correct_answers = "; ".join([a.text for a in answers if a.is_correct])

        ws.cell(row=row, column=1, value=question.id).border = thin_border
        ws.cell(row=row, column=2, value=question.text).border = thin_border
        ws.cell(row=row, column=3, value=question.question_type.name if question.question_type else "").border = thin_border
        ws.cell(row=row, column=4, value=all_answers).border = thin_border
        ws.cell(row=row, column=5, value=correct_answers).border = thin_border

        # Вставка изображений по столбцам (начиная с 6-й колонки)
        for i, image_obj in enumerate(question.images.all()):
            if i >= max_images:
                break  # на случай если что-то пошло не так
            image_path = os.path.join(settings.MEDIA_ROOT, image_obj.image.name)
            if os.path.exists(image_path):
                img = ExcelImage(image_path)
                img.width = 90
                img.height = 60
                col_letter = openpyxl.utils.get_column_letter(6 + i)
                cell_position = f"{col_letter}{row}"
                ws.add_image(img, cell_position)

                ws.cell(row=row, column=6 + i).border = thin_border
                ws.row_dimensions[row].height = 50
                ws.column_dimensions[col_letter].width = 15  # ширина под изображение

        row += 1

    # Автоширина для первых столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=questions.xlsx'
    wb.save(response)
    return response